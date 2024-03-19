from flask import Flask, request, jsonify 
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from openpyxl import load_workbook

# Create a Flask application
app = Flask(__name__)

# Configure the SQLAlchemy database connection
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///data.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Create instances of SQLAlchemy and Marshmallow
db = SQLAlchemy(app)
ma = Marshmallow(app)

# Define a model for the data I want to store in the database
class Parser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable = False)
    age = db.Column(db.String(50), nullable = False)
    city = db.Column(db.String(200), nullable = False)

    def __inti__(self, name, age, city):
        self.name = name
        self.age = age
        self.city = city

# Create a schema for the Parser model using Marshmallow
class ParserSchema(ma.Schema):
    class Meta:
        fields = ['id','name', 'age', 'city']

parser_schema = ParserSchema()
parser_schemas = ParserSchema(many=True)

# add data to the database from an Excel file
@app.route('/add', methods=['POST'])
def add_data():
    if request.method == 'POST':

        # Access the uploaded file from the request
        excel_data = request.files['Mydata']

        # Load the Excel workbook and get the active sheet
        Mydata = load_workbook(excel_data)
        Newdata = Mydata.active 

        # iter_rows is used to iterate over the rows starting from the second row 
        # skips the first row (row with index 1) because it is often assumed that the first row in an Excel sheet might contain headers or titles rather than data.
        for i in Newdata.iter_rows(min_row = 2, values_only = True): # values_only=True argument indicates that it should only return cell values, not cell objects.
            data = Parser(name=i[0], age=i[1], city=i[2])
            # data = Parser(**dict(zip(["name", "age", "city"], i))) we can also use this 
            # For each row, it creates a new Parser object by extracting values from the cells in that row. 
            db.session.add(data)
        db.session.commit()

    return "message: Data retrieve" 

# Get all data
@app.route('/get', methods=['GET'])
def get_all_data():
    # Query all records from the Parser model
    all_posts = Parser.query.all()

    # Serialize the data using the ParserSchema
    result = parser_schemas.dump(all_posts)

    # Return the serialized data as JSON
    return jsonify(result)

# Get single data
@app.route('/get/<int:id>', methods=['GET'])
def get_data(id):
    post = Parser.query.filter_by(id=id).first()
    result = parser_schema.dump(post)
    return jsonify(result)

# Update data
@app.route('/update/<int:id>', methods=['PUT'])
def update_data(id):
    post = Parser.query.get(id)

    # Extract data from the JSON request
    name = request.json['name']
    age = request.json['age']
    city = request.json['city']

    # Update the record with the new data
    post.name = name
    post.age = age
    post.city = city
    db.session.commit()
    return parser_schema.jsonify(post)

# Delete data
@app.route('/delete/<int:id>', methods=['DELETE'])
def delete_data(id):
    post = Parser.query.get(id)
    db.session.delete(post)
    db.session.commit()
    return parser_schema.jsonify(post)

if __name__ == '__main__':
    app.run(debug=True)

